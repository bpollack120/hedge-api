import pandas as pd
import requests
import datetime
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ORATS_TOKEN = "2749f06d-44ce-4b4c-8c39-46b5f1529a3e"

def get_unadjusted_daily_prices(ticker):
    url = "https://api.orats.io/datav2/hist/dailies"
    params = {"token": ORATS_TOKEN, "ticker": ticker}
    response = requests.get(url, params=params)
    if response.status_code != 200:
        print(f"Error fetching prices for {ticker}: {response.text}")
        return pd.DataFrame()
    try:
        data = response.json()
        df = pd.DataFrame(data["data"])
        df["date"] = pd.to_datetime(df["tradeDate"])
        df.set_index("date", inplace=True)
        df = df[["unadjClsPx"]].rename(columns={"unadjClsPx": "Price"})
        df.sort_index(inplace=True)
        return df
    except Exception as e:
        print("Error parsing price data:", e)
        return pd.DataFrame()

def get_next_friday(trade_date):
    days_until_friday = (4 - trade_date.weekday()) % 7
    return trade_date + datetime.timedelta(days=days_until_friday or 7)

def get_put_price_and_iv(ticker, trade_date, strike, expiry):
    url = "https://api.orats.io/datav2/hist/strikes/options"
    params = {
        "token": ORATS_TOKEN,
        "ticker": ticker,
        "tradeDate": trade_date.strftime("%Y-%m-%d"),
        "expirDate": expiry.strftime("%Y-%m-%d"),
        "strike": strike
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        try:
            data = response.json()
            for option in data.get("data", []):
                if option.get("putValue") is not None:
                    return option["putValue"], option.get("iv", None)
        except Exception as e:
            print(f"Error parsing options data: {e}")
    return None, None

def run_dynamic_hedge_analysis(ticker, start_date, end_date):
    prices = get_unadjusted_daily_prices(ticker)
    if prices.empty or not isinstance(prices.index, pd.DatetimeIndex):
        print(f"Aborting: No valid price data returned for {ticker}.")
        return

    prices = prices[start_date:end_date].copy()

    records = []
    current_position = None

    for date, row in prices.iterrows():
        spot = row["Price"]
        expiry = get_next_friday(date)
        target_strike = round(0.9 * spot / 5) * 5

        put_pnl = 0
        put_sale_price = None
        put_sale_strike = None
        rollover_flag = False

        # Rollover check
        if current_position:
            expiry_dt = pd.to_datetime(current_position["expiry"])
            # Rollover if strike or expiry changes
            if expiry != expiry_dt or target_strike != current_position["strike"]:
                put_sale_price, _ = get_put_price_and_iv(ticker, date, current_position["strike"], expiry_dt)
                if put_sale_price is not None:
                    put_pnl = (put_sale_price - current_position["entry_price"]) * 100
                    put_sale_strike = current_position["strike"]
                    rollover_flag = True
                current_position = None

        # Open new put (or carry forward if not rolled)
        put_price, iv = get_put_price_and_iv(ticker, date, target_strike, expiry)
        if put_price is None:
            continue

        stock_change = (spot - records[-1]["Price"]) * 100 if records else 0
        hedged_pnl = stock_change + put_pnl

        if current_position is None:
            current_position = {
                "strike": target_strike,
                "expiry": expiry,
                "entry_price": put_price
            }

        records.append({
            "Date": date,
            "Price": spot,
            "TargetStrike": target_strike,
            "Strike": current_position["strike"],
            "PutExpiry": current_position["expiry"].strftime("%Y-%m-%d"),
            "PutPrice": put_price,
            "PutIV": iv,
            "PutPnL": put_pnl,
            "PutSalePrice": put_sale_price,
            "PutSaleStrike": put_sale_strike,
            "StockChange": stock_change,
            "HedgedPnL": hedged_pnl,
            "UnhedgedPnL": stock_change,
            "Rollover": rollover_flag
        })

    df = pd.DataFrame(records)
    df["Hedged_Value"] = 100 * df["Price"].iloc[0] + df["HedgedPnL"].cumsum()
    df["Unhedged_Value"] = 100 * df["Price"].iloc[0] + df["UnhedgedPnL"].cumsum()

    # Output summary
    print(f"\nFinal Portfolio Values for {ticker}:")
    print(f"Unhedged: {df['Unhedged_Value'].iloc[-1]:.2f}")
    print(f"Hedged:   {df['Hedged_Value'].iloc[-1]:.2f}")
    initial_value = 100 * df["Price"].iloc[0]
    print(f"\nTotal Returns:")
    print(f"Unhedged: {(df['Unhedged_Value'].iloc[-1] - initial_value)/initial_value*100:.2f}%")
    print(f"Hedged:   {(df['Hedged_Value'].iloc[-1] - initial_value)/initial_value*100:.2f}%")

    # Save Excel
    os.makedirs("excel_exports", exist_ok=True)
    outname = os.path.join("excel_exports", f"{ticker}_dynamic_hedge_output.xlsx")
    df.to_excel(outname, index=False)

    # Highlight rollovers
    wb = load_workbook(outname)
    ws = wb.active
    fill = PatternFill(start_color="FFF3C9", end_color="FFF3C9", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        if ws[f"N{row}"].value is True:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
    wb.save(outname)

    # Save chart
    plt.figure(figsize=(12, 6))
    plt.title(f"{ticker}: Daily Dynamic Hedging with 10% OTM Put")
    plt.plot(df["Date"], df["Hedged_Value"], label="Hedged", linewidth=2)
    plt.plot(df["Date"], df["Unhedged_Value"], label="Unhedged", linestyle="--")
    plt.ylabel("Portfolio Value")
    plt.legend(loc="upper left")
    ax2 = plt.gca().twinx()
    ax2.plot(df["Date"], df["PutIV"], color='gray', alpha=0.4, label="Put IV")
    ax2.set_ylabel("Implied Volatility")
    plt.tight_layout()
    plt.savefig(os.path.join("excel_exports", f"{ticker}_dynamic_hedge_plot.png"))
    plt.close()

# Run it
if __name__ == "__main__":
    run_dynamic_hedge_analysis("TSLA", "2025-02-01", "2025-04-30")

